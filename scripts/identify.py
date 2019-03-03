import os
from openpyxl import load_workbook


def get_folder_names(path):
    """
    Will identify all of the folders in a given path and will return a list containing all the names by checking all of
    the available files/folders using isdir()

    Args:
        path - full path to a directory
    Returns:
        list of all subdirectories
    """
    folder_names = []
    for entry_name in os.listdir(path):
        entry_path = os.path.join(path, entry_name)
        if os.path.isdir(entry_path):
            folder_names.append(entry_name)
    return folder_names


# Will return all of the extensions present in the folders/sub-folders from the give path only once
def get_file_exts(path):
    file_exts = []
    list_files = os.walk(path)
    for root, _, files in list_files:
        for file_name in files:
            name, ext = os.path.splitext(file_name)
            ext = ext[1:].lower()  # in case the file is .PNG -> 'PNG' -> 'png'
            if ext not in file_exts:
                file_exts.append(ext)
    return file_exts


# Will count the num of files found in a directory and its subdirectories
def count_files(path):
    cpt = sum([len(files) for r, d, files in os.walk(path)])
    return cpt


# Find all the files from :path: and write them to :path_file: in a column :col:
def save_files(path, path_file, col):
    wb = load_workbook(path_file)
    ws = wb.active
    file_list = [files for r, d, files in os.walk(path)]
    flat_list = [item for sublist in file_list for item in sublist]
    i = 2  # Starting cell
    for item in flat_list:
        ws.cell(row=i, column=col).value = item
        i += 1
    wb.save(path_file)
